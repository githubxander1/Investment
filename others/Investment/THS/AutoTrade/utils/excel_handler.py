import os
from pprint import pprint

import openpyxl
import pandas as pd

from others.Investment.THS.AutoTrade.utils.scheduler import logger

def create_empty_excel(file_path, sheet_name):
    if not os.path.exists(file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            pd.DataFrame(columns=['组合名称', '代码', '操作', '新比例%', '时间']).to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info(f"创建空Excel文件: {file_path}, 表名称: {sheet_name}")

def read_excel(file_path, sheet_name):
    try:
        # 读取Excel文件
        if os.path.exists(file_path):
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            # pprint(f"成功读取文件: {file_path}, 表名称: {sheet_name}")
            return df
        else:
            logger.warning(f"文件 {file_path} 不存在，返回空DataFrame")
            return pd.DataFrame()
    except Exception as e:
        logger.error(f"读取文件失败: {e}")
        return pd.DataFrame()

def save_to_excel(df, filename, sheet_name, mode='w', index=True):
    """增强版保存函数，支持索引持久化"""
    if mode == 'a' and os.path.exists(filename):
        book = openpyxl.load_workbook(filename)
        if sheet_name in book.sheetnames:
            start_idx = book[sheet_name].max_row
        else:
            start_idx = 0
            
        # 保留DataFrame索引
        df.index += start_idx
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            writer.book = book  # 保留现有工作簿
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df.to_excel(writer, sheet_name=sheet_name, startrow=start_idx, index=index, header=False)
    else:
        df.to_excel(filename, sheet_name=sheet_name, index=index)

def clear_sheet(filename, sheet_name):
    """清空指定Excel文件中的指定表格"""
    try:
        # 检查文件是否存在
        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 删除所有行
                ws.delete_rows(1, ws.max_row)
                wb.save(filename)
                pprint(f"成功清空表格: {sheet_name} 文件: {filename}")
            else:
                logger.warning(f"表格 {sheet_name} 不存在于文件: {filename}")
        else:
            logger.warning(f"文件 {filename} 不存在，无需清空")
    except Exception as e:
        logger.error(f"清空表格失败: {e}")
