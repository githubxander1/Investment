# scripts/策略_今天调仓.py
import datetime
import os

# import UserAgent
import openpyxl
import pandas as pd
import requests
from fake_useragent import UserAgent

from others.Investment.THS.AutoTrade.config.settings import STRATEGY_TODAY_ADJUSTMENT_LOG_FILE, \
    STRATEGY_TODAY_ADJUSTMENT_FILE, OPRATION_RECORD_DONE_FILE, Strategy_id_to_name, Strategy_ids, \
    ETF_Combination_TODAY_ADJUSTMENT_FILE
from others.Investment.THS.AutoTrade.utils.api_client import APIClient
# from others.Investment.THS.AutoTrade.utils.api_client import APIClient
# from utils.api_client import APIClient
from others.Investment.THS.AutoTrade.utils.determine_market import determine_market
# from utils.determine_market import determine_market
from others.Investment.THS.AutoTrade.utils.logger import setup_logger
# from utils.logger import setup_logger
from others.Investment.THS.AutoTrade.utils.notification import send_notification, send_email

# from utils.notification import send_notification

# from utils.notification import

logger = setup_logger(STRATEGY_TODAY_ADJUSTMENT_LOG_FILE)

api_client = APIClient()

ua = UserAgent()

# 手动创建策略ID到策略名称的映射


def send_http_request(url, params, headers):
    try:
        response = requests.get(url, params=params, headers=headers)
        response.raise_for_status()  # 检查请求是否成功
        return response.json()
    except requests.RequestException as e:
        logger.error(f"请求失败: {e}")
        return None

def get_latest_position_and_trade(strategy_id):
    strategy_id_to_name = Strategy_id_to_name
    """获取策略的最新持仓和交易信息"""
    url = "https://ms.10jqka.com.cn/iwencai/iwc-web-business-center/strategy_unify/strategy_profit"
    params = {"strategyId": strategy_id}

    headers = {
        "User-Agent": ua.random
    }

    data = send_http_request(url, params, headers)
    # pprint(data)
    if data:
        latest_trade = data.get('result', {}).get('latestTrade', {})
        trade_date = latest_trade.get('tradeDate', 'N/A')
        trade_stocks = latest_trade.get('tradeStocks', [])

        latest_trade_info = []
        for trade_info in trade_stocks:
            code = trade_info.get('stkCode', 'N/A').split('.')[0]
            trade_entry = {
                '策略名称': strategy_id_to_name.get(strategy_id, '未知策略'),
                '时间': trade_date,
                '操作': trade_info.get('operationType', 'N/A'),
                '市场': determine_market(code),
                '股票名称': trade_info.get('stkName', 'N/A'),
                '最新价': trade_info.get('tradePrice', 'N/A'),
                '新比例%': round(trade_info.get('position', 'N/A') * 100,2)
            }
            latest_trade_info.append(trade_entry)
        return latest_trade_info, trade_date
    else:
        return [], 'N/A'

def save_to_excel(df, filename, sheet_name, index=False):
    """将DataFrame保存到Excel文件中"""
    try:
        # 检查文件是否存在
        try:
            with pd.ExcelFile(filename) as _:
                # 文件存在，追加模式
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=index)
        except FileNotFoundError:
            # 文件不存在，创建新文件
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
        logger.info(f"成功保存数据到文件: {filename}, 表名称: {sheet_name}")
    except Exception as e:
        logger.error(f"保存数据到文件失败: {e}")

def clear_sheet(filename, sheet_name):
    """清空指定Excel文件中的指定表格"""
    try:
        # 检查文件是否存在
        try:
            wb = openpyxl.load_workbook(filename)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 删除所有行
                ws.delete_rows(1, ws.max_row)
                wb.save(filename)
                logger.info(f"成功清空表格: {sheet_name} 文件: {filename}")
            else:
                logger.warning(f"表格 {sheet_name} 不存在于文件: {filename}")
        except FileNotFoundError:
            logger.warning(f"文件 {filename} 不存在，无需清空")
    except Exception as e:
        logger.error(f"清空表格失败: {e}")

async def check_new_data(existing_df, today_trade_df, sheet_name):
    """优化后的数据合并函数

    对比新旧数据
    找出新增
    附加补充新增数据到旧表

    Args:
        existing_df: 已存在的历史数据DataFrame
        today_trade_df: 当天获取的新数据DataFrame
        sheet_name: 要操作的Excel表名称

    Returns:
        pd.DataFrame: 合并后的完整数据集
    """
    try:
        # 预处理时间字段（精确到分钟）
        today_trade_df = today_trade_df.copy()
        today_trade_df['时间'] = pd.to_datetime(today_trade_df['时间']).dt.floor('Min')

        # 生成唯一标识（组合名称+代码+操作+新比例+时分）
        def create_unique_id(row):
            try:
                # 统一股票代码为6位格式（自动补零）
                formatted_code = str(row['代码']).zfill(6)  # 关键修复点

                ratio = "{:.2f}".format(float(row['新比例%']))
                time_str = row['时间'].strftime('%H%M%S')
                return f"{row['组合名称']}_{formatted_code}_{row['操作']}_{ratio}_{time_str}"
            except Exception as e:
                logger.error(f"生成唯一标识失败: {str(e)}")
                return "error_id"


        today_trade_df['strict_id'] = today_trade_df.apply(create_unique_id, axis=1)

        # 处理空历史数据情况
        if existing_df.empty:
            # 首次保存时添加时间排序
            sorted_df = today_trade_df.sort_values('时间', ascending=True)
            save_to_excel(sorted_df, ETF_Combination_TODAY_ADJUSTMENT_FILE, sheet_name)
            # # 原代码中的通知部分：
            # send_notification(f"首次发现调仓，{sheet_name}")
            #
            # # 修改为：
            # notification_msg = f"{sheet_name}调仓操作\n" + "\n".join(
            #     [f"{row['组合名称']} {row['操作']} {row['代码']}"
            #      for _, row in new_data.iterrows()])
            # send_notification(notification_msg)

            return sorted_df

        # 标准化历史数据格式
        existing_df = existing_df.copy()
        existing_df['strict_id'] = existing_df.apply(create_unique_id, axis=1)  # 统一列名

        # 找出新增数据（修复查询条件）
        mask = ~today_trade_df['strict_id'].isin(existing_df['strict_id'])
        new_data = today_trade_df[mask].copy()

        # 合并数据并去重
        if not new_data.empty:
            # 合并前再次去重（防止多次运行产生重复）
            updated_df = pd.concat([existing_df, new_data], ignore_index=True)
            updated_df = updated_df.drop_duplicates(subset=['strict_id'], keep='last')

            # 按精确时间排序（秒级精度）
            sorted_df = updated_df.sort_values('时间', ascending=True).reset_index(drop=True)

            # 原子写入操作
            temp_file = ETF_Combination_TODAY_ADJUSTMENT_FILE.replace('.xlsx', '_temp.xlsx')
            save_to_excel(sorted_df, temp_file, sheet_name)
            # send_notification(f"新调仓，{sheet_name}")
            # # 原代码中的通知部分：
            # send_notification(f"首次发现调仓，{sheet_name}")

            # # 修改为：{sheet_name}操作
            notification_msg = f"\n> " + "\n".join(
                [f"{row['策略名称']} {row['操作']} {row['股票名称']} {row['新比例%']}% {row['最新价']} \n{row['时间']}"
                 for _, row in new_data.iterrows()])
            send_notification(notification_msg)

            if os.path.exists(temp_file):
                os.replace(temp_file, ETF_Combination_TODAY_ADJUSTMENT_FILE)

            logger.info(f"新增{len(new_data)}条唯一调仓记录")
        else:
            logger.info("没有新增调仓数据")
            sorted_df = existing_df

        return sorted_df.drop(columns=['strict_id'], errors='ignore')

    except Exception as e:
        logger.error(f"数据处理失败: {str(e)}")
        # 保留临时文件供调试
        if 'temp_file' in locals() and os.path.exists(temp_file):
            logger.error(f"临时文件保留在: {temp_file}")
        return existing_df
async def strategy_main():
    strategy_id_to_name = Strategy_id_to_name
    strategy_ids = Strategy_ids
    all_today_trades_info = []
    all_latest_trade_info = []

    logger.info("开始处理策略调仓信息")
    for strategy_id in strategy_ids:
        combination_name = Strategy_id_to_name.get(strategy_id, '未知策略')
        latest_trade_info, trade_date = get_latest_position_and_trade(strategy_id)
        if latest_trade_info:
            all_latest_trade_info.extend(latest_trade_info)
            today_trades = [trade for trade in latest_trade_info if trade['时间'] == datetime.datetime.now().date().strftime('%Y%m%d')]
            all_today_trades_info.extend(today_trades)

    # 过滤掉创业板股票的交易信息
    all_latest_trade_info_without_sc = [trade for trade in all_latest_trade_info if trade['市场'] not in ['创业板', '科创板']]
    all_today_trades_info_without_sc = [trade for trade in all_today_trades_info if trade['市场'] not in ['创业板', '科创板']]
    # logger.info("去掉参考价大于30的")
    # all_today_trades_info_without_sc = [trade for trade in all_today_trades_info_without_sc if trade['参考价'] < 30]

    today_trades_without_sc_df = pd.DataFrame(all_today_trades_info_without_sc)

    file_path = STRATEGY_TODAY_ADJUSTMENT_FILE

    # 清空昨天的数据
    clear_sheet(file_path, '策略今天调仓')  # 如果不需要清空，可以注释掉

    # 保存今天的数据，即使为空也会覆盖昨天的数据
    save_to_excel(today_trades_without_sc_df, file_path, sheet_name='策略今天调仓', index=False)
    logger.info(f'今日调仓：\n {today_trades_without_sc_df}')

    if not today_trades_without_sc_df.empty:
        # 发送通知
        # send_notification("今日有新的调仓操作！策略")
        # send_email('股票策略调仓', today_trades_without_sc_df.to_string(), '2695418206@qq.com')

        # 原代码中的通知部分：
        # send_notification(f"首次发现调仓，{sheet_name}")

        # 修改为：
        notification_msg = f"\n" + "\n".join(
            [f"{row['策略名称']} {row['操作']} {row['股票名称']} {row['新比例%']}% {row['最新价']}"
             for _, row in today_trades_without_sc_df.iterrows()])
        send_notification(notification_msg)

        # 创建标志文件
        with open(f"{OPRATION_RECORD_DONE_FILE}", "w") as f:
            f.write("策略调仓已完成")
            logger.info("创建标志文件成功")
    else:
        logger.info("未发送通知: 策略今天有调仓，但是是非沪深股票或无新增调仓")

    logger.info("策略调仓信息处理完成")

if __name__ == '__main__':
    strategy_id_to_name = Strategy_id_to_name
    import asyncio

    asyncio.run(strategy_main())

    # try:
    #     scheduler = Scheduler(interval=0.25,
    #                           callback=main,
    #                           start_time=dt_time(9, 29),
    #                           end_time=dt_time(10, 33))
    #     scheduler.start()
    # except Exception as e:
    #     logger.error(f"调度器启动失败: {e}", exc_info=True)

