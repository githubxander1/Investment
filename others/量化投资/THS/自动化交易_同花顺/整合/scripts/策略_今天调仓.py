import logging
import datetime
from pprint import pprint

import openpyxl
import pandas as pd

import requests
from fake_useragent import UserAgent

from others.量化投资.THS.自动化交易_同花顺.整合.utils.api_client import APIClient
from others.量化投资.THS.自动化交易_同花顺.整合.utils.determine_market import determine_market
from others.量化投资.THS.自动化交易_同花顺.整合.utils.notification import send_notification
from others.量化投资.THS.自动化交易_同花顺.整合.config.settings import STRATEGY_TODAY_ADJUSTMENT_LOG_FILE, API_URL, HEADERS, \
    STRATEGY_TODAY_ADJUSTMENT_FILE, STRATEGY_ID_TO_NAME, OPRATION_RECORD_DONE_FILE
from others.量化投资.THS.自动化交易_同花顺.整合.utils.ths_logger import setup_logger

logger = setup_logger(STRATEGY_TODAY_ADJUSTMENT_LOG_FILE)

api_client = APIClient()

ua = UserAgent()

# 手动创建策略ID到策略名称的映射


def get_latest_position_and_trade(strategy_id):
    """获取策略的最新持仓和交易信息"""
    url = "https://ms.10jqka.com.cn/iwencai/iwc-web-business-center/strategy_unify/strategy_profit"
    params = {"strategyId": strategy_id}

    headers = {
        "User-Agent": ua.random
    }

    try:
        response = requests.get(url, params=params, headers=headers)
        response.raise_for_status()  # 检查请求是否成功
        data = response.json()
    except requests.RequestException as e:
        logger.error(f"请求失败: {e}")
        return [], 'N/A'

    if data:
        latest_trade = data.get('result', {}).get('latestTrade', {})
        trade_date = latest_trade.get('tradeDate', 'N/A')
        trade_stocks = latest_trade.get('tradeStocks', [])

        latest_trade_info = []
        for trade_info in trade_stocks:
            code = trade_info.get('stkCode', 'N/A').split('.')[0]
            trade_entry = {
                '策略名称': strategy_id_to_name.get(strategy_id, '未知策略'),
                '时间': trade_date,
                '操作': trade_info.get('operationType', 'N/A'),
                '市场': determine_market(code),
                '股票名称': trade_info.get('stkName', 'N/A'),
                '参考价': trade_info.get('tradePrice', 'N/A'),
                '数量': trade_info.get('tradeAmount', 'N/A'),
            }
            latest_trade_info.append(trade_entry)
        return latest_trade_info, trade_date
    else:
        return [], 'N/A'

def save_to_excel(df, filename, sheet_name, index=False):
    """将DataFrame保存到Excel文件中"""
    try:
        # 检查文件是否存在
        try:
            with pd.ExcelFile(filename) as _:
                # 文件存在，追加模式
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=index)
        except FileNotFoundError:
            # 文件不存在，创建新文件
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
        logger.info(f"成功保存数据到文件: {filename}, 表名称: {sheet_name}")
    except Exception as e:
        logger.error(f"保存数据到文件失败: {e}")

def clear_sheet(filename, sheet_name):
    """清空指定Excel文件中的指定表格"""
    try:
        # 检查文件是否存在
        try:
            wb = openpyxl.load_workbook(filename)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 删除所有行
                ws.delete_rows(1, ws.max_row)
                wb.save(filename)
                logger.info(f"成功清空表格: {sheet_name} 文件: {filename}")
            else:
                logger.warning(f"表格 {sheet_name} 不存在于文件: {filename}")
        except FileNotFoundError:
            logger.warning(f"文件 {filename} 不存在，无需清空")
    except Exception as e:
        logger.error(f"清空表格失败: {e}")

def main():
    strategy_ids = ['138006', '137789', '155259', '155270', '155680']
    all_today_trades_info = []
    all_latest_trade_info = []

    logger.info("开始处理策略调仓信息")
    for strategy_id in strategy_ids:
        combination_name = STRATEGY_ID_TO_NAME.get(strategy_id, '未知策略')
        latest_trade_info, trade_date = get_latest_position_and_trade(strategy_id)
        if latest_trade_info:
            all_latest_trade_info.extend(latest_trade_info)
            today_trades = [trade for trade in latest_trade_info if trade['时间'] == datetime.datetime.now().date().strftime('%Y%m%d')]
            all_today_trades_info.extend(today_trades)

    # 过滤掉创业板股票的交易信息
    all_latest_trade_info_without_sc = [trade for trade in all_latest_trade_info if trade['市场'] not in ['创业板', '科创板']]
    all_today_trades_info_without_sc = [trade for trade in all_today_trades_info if trade['市场'] not in ['创业板', '科创板']]

    today_trades_without_sc_df = pd.DataFrame(all_today_trades_info_without_sc)

    file_path = STRATEGY_TODAY_ADJUSTMENT_FILE

    # 清空昨天的数据
    clear_sheet(file_path, '策略今天调仓')  # 如果不需要清空，可以注释掉

    # 保存今天的数据，即使为空也会覆盖昨天的数据
    save_to_excel(today_trades_without_sc_df, file_path, sheet_name='策略今天调仓', index=False)
    logger.info(f'今日调仓：\n {today_trades_without_sc_df}')

    if not today_trades_without_sc_df.empty:
        # 发送通知
        send_notification("今日有新的调仓操作！策略")
        # 创建标志文件
        with open(f"{OPRATION_RECORD_DONE_FILE}", "w") as f:
            f.write("策略调仓已完成")
            logger.info("创建标志文件成功")
    else:
        logger.info("未发送通知: 策略今天有调仓，但是是非沪深股票或无新增调仓")

    logger.info("策略调仓信息处理完成")

if __name__ == '__main__':
    strategy_id_to_name = {
        '155259': 'TMT资金流入战法',
        '155680': 'GPT定期精选',
        '138036': '低价小盘股战法',
        '155270': '中字头概念',
        '137789': '高现金毛利战法',
        '138006': '连续五年优质股战法',
        '136567': '净利润同比大增低估值战法',
        '138127': '归母净利润高战法',
        '118188': '均线粘合平台突破'
    }
    main()
