# scripts/组合_今天调仓.py
import datetime
import os
from datetime import time

import openpyxl
import pandas as pd
import requests
from fake_useragent import UserAgent

from others.量化投资.THS.自动化交易_同花顺.整合.config.settings import COMBINATION_TODAY_ADJUSTMENT_LOG_FILE, \
    COMBINATION_TODAY_ADJUSTMENT_FILE, OPRATION_RECORD_DONE_FILE
from others.量化投资.THS.自动化交易_同花顺.整合.utils.determine_market import determine_market
from others.量化投资.THS.自动化交易_同花顺.整合.utils.notification import send_notification
from others.量化投资.THS.自动化交易_同花顺.整合.utils.scheduler import Scheduler
from others.量化投资.THS.自动化交易_同花顺.整合.utils.ths_logger import setup_logger

logger = setup_logger(COMBINATION_TODAY_ADJUSTMENT_LOG_FILE)

ua = UserAgent()

def get_strategy_details(product_id):
    """获取组合名称和描述"""
    url = "https://dq.10jqka.com.cn/fuyao/tg_package/package/v1/get_package_portfolio_infos"
    headers = {
        "User-Agent": "Mozilla/5.0 (Linux; Android 10; Redmi Note 7 Pro Build/QKQ1.190915.002; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/87.0.4280.101 Mobile Safari/537.36 Hexin_Gphone/11.16.10 (Royal Flush) hxtheme/1 innerversion/G037.08.980.1.32 followPhoneSystemTheme/1 userid/641926488 getHXAPPAccessibilityMode/0 hxNewFont/1 isVip/0 getHXAPPFontSetting/normal getHXAPPAdaptOldSetting/0",
        # "Referer": "https://t.10jqka.com.cn/pkgfront/tgService.html?type=portfolio&id=19483",
        "Cookie": "IFUserCookieKey={}; user=MDptb182NDE5MjY0ODg6Ok5vbmU6NTAwOjY1MTkyNjQ4ODo3,ExMTExMTExMTExLDQwOzQ0LDExLDQwOzYsMSw0MDs1LDEsNDA7MSwxMDEsNDA7MiwxLDQwOzMsMSw0MDs1LDEsNDA7OCwwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMSw0MDsxMDIsMSw0MDoyNzo6OjY0MTkyNjQ4ODoxNzM0MDUzNTg5Ojo6MTY1ODE0Mjc4MDoyNjc4NDAwOjA6MTE3MTRjYTYwODhjNjRmYzZmNDFlZDRkOTJhMDU3NTMwOjox; userid=641926488; u_name=mo_641926488; escapename=mo_641926488; ticket=58d0f4bf66d65411bb8d8aa431e00721; user_status=0; hxmPid=sns_my_pay_new; v=AxLXmrX7ofaqkd2K73acRpPBYdP0Ixa9SCcK4dxrPkWw771JxLNmzRi3WvOv"
    }
    params = {"product_id": product_id, "product_type": "portfolio"}
    try:
        response = requests.get(url, params=params, headers=headers)
        response.raise_for_status()
        result = response.json()
        if result['status_code'] == 0:
            return {
                "组合id": product_id,
                "组合名称": result['data']['baseInfo']['productName'],
                "组合描述": result['data']['baseInfo']['productDesc']
            }
        else:
            logger.error(f"Failed to retrieve data for product_id: {product_id}")
            return None
    except requests.RequestException as e:
        logger.error(f"请求出现错误: {e}")
        return None

def get_historical_data(portfolio_id):
    """获取历史调仓数据"""
    url = "https://t.10jqka.com.cn/portfolio/post/v2/get_relocate_post_list"

    headers = {
        "User-Agent": "Mozilla/5.0 (Linux; Android 10; Redmi Note 7 Pro Build/QKQ1.190915.002; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/87.0.4280.101 Mobile Safari/537.36 Hexin_Gphone/11.16.10 (Royal Flush) hxtheme/1 innerversion/G037.08.980.1.32 followPhoneSystemTheme/1 userid/641926488 getHXAPPAccessibilityMode/0 hxNewFont/1 isVip/0 getHXAPPFontSetting/normal getHXAPPAdaptOldSetting/0",
        "Cookie": "IFUserCookieKey={}; user=MDptb182NDE5MjY0ODg6Ok5vbmU6NTAwOjY1MTkyNjQ4ODo3LDExMTExMTExMTExLDQwOzQ0LDExLDQwOzYsMSw0MDs1LDEsNDA7MSwxMDEsNDA7MiwxLDQwOzMsMSw0MDs1LDEsNDA7OCwwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMSw0MDsxMDIsMSw0MDoyNzo6OjY0MTkyNjQ4ODoxNzM0MDUzNTg5Ojo6MTY1ODE0Mjc4MDoyNjc4NDAwOjA6MTE3MTRjYTYwODhjNjRmYzZmNDFlZDRkOTJhMDU3NTMwOjox; userid=641926488; u_name=mo_641926488; escapename=mo_641926488; ticket=58d0f4bf66d65411bb8d8aa431e00721; user_status=0; hxmPid=sns_my_pay_new; v=A-gtFDtpG9AGTjdUiWYWoHVzu936EUwbLnUgn6IZNGNW_YfHSiEcq36F8Czx"
    }
    params = {"id": portfolio_id, "dynamic_id": 0}
    try:
        response = requests.get(url, params=params, headers=headers)
        response.raise_for_status()
        # pprint(response.json())
        return response.json()
    except requests.RequestException as e:
        logger.error(f"请求出现错误: {e}")
        return None

def process_today_trades(ids):
    """处理今天调仓数据"""
    all_records = []
    today = datetime.date.today().strftime('%Y-%m-%d')
    # print(f'今天{today}')
    # processed_ids = set(read_processed_ids())

    for portfolio_id in ids:
        data = get_historical_data(portfolio_id)
        # pprint(data)
        if data:
            for item in data['data']:
                create_at = item['createAt']
                # print(f'时间{create_at}')
                date_part = create_at.split()[0]
                if date_part == today:
                    content = item['content']
                    # need_relocate_reason = item['needRelocateReason']
                    for relocate in item['relocateList']:
                        code = relocate['code']
                        current_ratio = relocate['currentRatio']
                        final_price = relocate['finalPrice']
                        name = relocate['name']
                        # print(name)

                        # 检查股票名称是否为“匿名”
                        if '***' in name:
                            logger.warning(f"未订阅或股票名称显示异常 - 股票代码: {code}, 时间: {create_at}")
                            continue

                        new_ratio = relocate['newRatio']
                        market = determine_market(code)
                        operation = 'SALE' if new_ratio < current_ratio else 'BUY'
                        all_records.append({
                            '组合id': portfolio_id,
                            '组合名称': get_strategy_details(portfolio_id).get("组合名称"),
                            '描述': get_strategy_details(portfolio_id).get("组合描述"),
                            '说明': content,
                            '时间': pd.to_datetime(create_at),
                            '操作': operation,
                            '市场': market,
                            '股票名称': name,
                            '参考价': final_price,
                            '当前比例': f"{current_ratio * 100:.2f}%",
                            '新比例': f"{new_ratio * 100:.2f}%"
                        })
                            # processed_ids.add(unique_id)
    # pprint(all_records)
    if not all_records:
        logger.info("所选组合今天无调仓")
        return None

    today_trade_df = pd.DataFrame(all_records)

    today_trade_print_df = today_trade_df.drop(columns=['组合id', '描述', '说明'])
    today_trade_without_cyb_print_df = today_trade_print_df[~today_trade_df['市场'].isin(['创业板', '科创板'])]
    logger.info('去掉创业板和科创板')

    return today_trade_without_cyb_print_df

def save_to_excel(df, filename, sheet_name, index=False):
    """追加保存DataFrame到Excel文件"""
    try:
        # 检查文件是否存在
        try:
            with pd.ExcelFile(filename) as _:
                # 文件存在，追加模式
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    # 检查表是否存在
                    if sheet_name not in writer.book.sheetnames:
                        df.to_excel(writer, sheet_name=sheet_name, index=index)
                    else:
                        df.to_excel(writer, sheet_name=sheet_name, index=index)
                logger.info(f"成功追加数据到Excel文件: {filename}, 表名称: {sheet_name}")
        except FileNotFoundError:
            # 文件不存在，创建新文件
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index, header=True)
            logger.info(f"成功创建并保存数据到Excel文件: {filename}, 表名称: {sheet_name}")
    except Exception as e:
        logger.error(f"保存数据到Excel文件失败: {e}")

def clear_sheet(filename, sheet_name):
    """清空指定Excel文件中的指定表格"""
    try:
        # 检查文件是否存在
        try:
            wb = openpyxl.load_workbook(filename)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 删除所有行
                ws.delete_rows(1, ws.max_row)
                wb.save(filename)
                logger.info(f"成功清空表格: {sheet_name} 在文件: {filename}")
            else:
                logger.warning(f"表格 {sheet_name} 不存在于文件: {filename}")
        except FileNotFoundError:
            logger.warning(f"文件 {filename} 不存在，无需清空")
    except Exception as e:
        logger.error(f"清空表格失败: {e}")

def create_flag_file(flag_file):
    """创建标志文件"""
    try:
        with open(flag_file, "w") as f:
            f.write("已创建标志文件")
        logger.info(f"创建标志文件: {flag_file}")
    except Exception as e:
        logger.error(f"创建标志文件失败: {e}")

def check_clear_flag(flag_file):
    if not os.path.exists(flag_file):
        # 清空昨天的数据
        try:
            clear_sheet(file_path, '组合今天调仓')
            # logger.info("清空昨天的数据")
            # 创建清空标志文件
            create_flag_file(flag_file)

        except Exception as e:
            logger.error(f"清空表格失败: {e}")
            return

def check_new_data(existing_df, today_trade_without_cyb_print_df):
    # 读取上次保存的结果,检查
    if not existing_df.empty:
        new_data = today_trade_without_cyb_print_df.merge(existing_df, how='outer', indicator=True).loc[
            lambda x: x['_merge'] == 'left_only'].drop(columns='_merge')

    else:
        new_data = today_trade_without_cyb_print_df.copy()


    if not new_data.empty:
        logger.info(f'新增调仓：\n {new_data}')
        # 合并新旧数据并排序
        combined_df = pd.concat([new_data, today_trade_without_cyb_print_df], ignore_index=True)
        #去重
        # combined_df.drop_duplicates(subset=['股票名称', '操作', '时间'], keep='first', inplace=True)
        # 去重，对比所有字段
        combined_df.drop_duplicates(subset=None, keep='first', inplace=True)
        combined_df['时间'] = pd.to_datetime(combined_df['时间'])  # 确保时间列是datetime类型
        combined_df.sort_values(by='时间', ascending=True, inplace=True)  # 按时间降序排列
        logger.info(f'合并新旧数据：\n {combined_df}')
        # print(f'合并新旧数据：\n {combined_df}')

        save_to_excel(combined_df, file_path, sheet_name='组合今天调仓', index=False)

        send_notification("今天有新调仓，组合")

        # 创建标志文件
        create_flag_file(OPRATION_RECORD_DONE_FILE)
    else:
        logger.info("没有新增调仓数据")

def main():
    combination_ids = [6994, 7152,18710, 16281, 19347, 13081, 14980]
    '''
    13081 好赛道出牛股
16281 每天进步一点点
18565 龙头一年三倍

7152 中线龙头
6994 梦想一号  死群，调仓好久好久甚至半年
11094 低位题材
14980 波段突击
19347 超短稳定复利
18710 用收益率征服您'''
    # combination_ids = ['14980']
    logger.info("开始处理组合调仓信息")
    today_trade_without_cyb_print_df = process_today_trades(combination_ids)

    # existing_df = pd.DataFrame()
    # existing_df.to_excel(file_path, sheet_name='组合今天调仓', index=False)
    if today_trade_without_cyb_print_df is not None:
        logger.info(f'今日调仓：\n {today_trade_without_cyb_print_df}')
        try:
            existing_df = pd.read_excel(file_path, sheet_name='组合今天调仓')
        except FileNotFoundError:
            existing_df = pd.DataFrame()
            existing_df.to_excel(file_path, sheet_name='组合今天调仓', index=False)
            logger.info("Excel文件不存在，创建新文件")

        check_new_data(existing_df, today_trade_without_cyb_print_df)

    else:
        logger.info("今天没有新的调仓操作")

    # logger.info("组合调仓信息处理完成")

if __name__ == '__main__':
    file_path = COMBINATION_TODAY_ADJUSTMENT_FILE
    main()

    start_time = time(9, 30)  # 九点半
    end_time = time(15, 00)    # 下午三点
    scheduler = Scheduler(1, main, start_time, end_time)
    try:
        scheduler.start()
    except Exception as e:
        logger.error(f"调度器启动失败: {e}", exc_info=True)
    # finally:
    #     if os.path.exists(CLEAR_FLAG_FILE):
    #         os.remove(CLEAR_FLAG_FILE)
    #         logger.info("删除清空标志文件")