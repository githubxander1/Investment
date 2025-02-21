# scripts/策略_今天调仓.py
import datetime

import openpyxl
import pandas as pd
import requests
from fake_useragent import UserAgent

# from config.settings import STRATEGY_TODAY_ADJUSTMENT_LOG_FILE
from others.量化投资.THS.自动化交易_同花顺.config.settings import STRATEGY_TODAY_ADJUSTMENT_LOG_FILE, \
    STRATEGY_TODAY_ADJUSTMENT_FILE, OPRATION_RECORD_DONE_FILE, Strategy_id_to_name, Strategy_ids
from others.量化投资.THS.自动化交易_同花顺.utils.api_client import APIClient
# from config.settings import STRATEGY_TODAY_ADJUSTMENT_LOG_FILE, \
#     STRATEGY_TODAY_ADJUSTMENT_FILE, OPRATION_RECORD_DONE_FILE, Strategy_id_to_name, Strategy_ids
# from others.量化投资.THS.自动化交易_同花顺.utils.api_client import APIClient
# from utils.api_client import APIClient
from others.量化投资.THS.自动化交易_同花顺.utils.determine_market import determine_market
# from utils.determine_market import determine_market
from others.量化投资.THS.自动化交易_同花顺.utils.logger import setup_logger
# from utils.logger import setup_logger
from others.量化投资.THS.自动化交易_同花顺.utils.notification import send_notification, send_email

# from utils.notification import send_notification

# from utils.notification import

logger = setup_logger(STRATEGY_TODAY_ADJUSTMENT_LOG_FILE)

api_client = APIClient()

ua = UserAgent()

# 手动创建策略ID到策略名称的映射


def send_http_request(url, params, headers):
    try:
        response = requests.get(url, params=params, headers=headers)
        response.raise_for_status()  # 检查请求是否成功
        return response.json()
    except requests.RequestException as e:
        logger.error(f"请求失败: {e}")
        return None

def get_latest_position_and_trade(strategy_id):
    strategy_id_to_name = Strategy_id_to_name
    """获取策略的最新持仓和交易信息"""
    url = "https://ms.10jqka.com.cn/iwencai/iwc-web-business-center/strategy_unify/strategy_profit"
    params = {"strategyId": strategy_id}

    headers = {
        "User-Agent": ua.random
    }

    data = send_http_request(url, params, headers)
    # pprint(data)
    if data:
        latest_trade = data.get('result', {}).get('latestTrade', {})
        trade_date = latest_trade.get('tradeDate', 'N/A')
        trade_stocks = latest_trade.get('tradeStocks', [])

        latest_trade_info = []
        for trade_info in trade_stocks:
            code = trade_info.get('stkCode', 'N/A').split('.')[0]
            trade_entry = {
                '策略名称': strategy_id_to_name.get(strategy_id, '未知策略'),
                '时间': trade_date,
                '操作': trade_info.get('operationType', 'N/A'),
                '市场': determine_market(code),
                '股票名称': trade_info.get('stkName', 'N/A'),
                '参考价': trade_info.get('tradePrice', 'N/A'),
                '新比例%': round(trade_info.get('position', 'N/A') * 100,2)
            }
            latest_trade_info.append(trade_entry)
        return latest_trade_info, trade_date
    else:
        return [], 'N/A'

def save_to_excel(df, filename, sheet_name, index=False):
    """将DataFrame保存到Excel文件中"""
    try:
        # 检查文件是否存在
        try:
            with pd.ExcelFile(filename) as _:
                # 文件存在，追加模式
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=index)
        except FileNotFoundError:
            # 文件不存在，创建新文件
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
        logger.info(f"成功保存数据到文件: {filename}, 表名称: {sheet_name}")
    except Exception as e:
        logger.error(f"保存数据到文件失败: {e}")

def clear_sheet(filename, sheet_name):
    """清空指定Excel文件中的指定表格"""
    try:
        # 检查文件是否存在
        try:
            wb = openpyxl.load_workbook(filename)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 删除所有行
                ws.delete_rows(1, ws.max_row)
                wb.save(filename)
                logger.info(f"成功清空表格: {sheet_name} 文件: {filename}")
            else:
                logger.warning(f"表格 {sheet_name} 不存在于文件: {filename}")
        except FileNotFoundError:
            logger.warning(f"文件 {filename} 不存在，无需清空")
    except Exception as e:
        logger.error(f"清空表格失败: {e}")

# def check_new_data(existing_df, today_trade_df, sheet_name):
#     # 确保 '代码' 列的数据类型一致，并统一格式化为6位字符串
#     today_trade_df['代码'] = today_trade_df['代码'].astype(str).str.zfill(6)
#     if not existing_df.empty:
#         existing_df['代码'] = existing_df['代码'].astype(str).str.zfill(6)
#
#     # 确保 '时间' 列的数据类型一致
#     today_trade_df['时间'] = pd.to_datetime(today_trade_df['时间'])
#     if not existing_df.empty:
#         existing_df['时间'] = pd.to_datetime(existing_df['时间'])
#
#     # 确保 '组合名称' 列的数据类型一致
#     today_trade_df['组合名称'] = today_trade_df['组合名称'].astype(str)
#     if not existing_df.empty:
#         existing_df['组合名称'] = existing_df['组合名称'].astype(str)
#
#     # 合并数据并标记新增数据
#     if existing_df.empty:
#         new_data = today_trade_df.copy()
#
#     # 找出df1中与df2不同的行
#     # merged = pd.merge(existing_df, today_trade_df, how='outer', indicator=True)
#     # different_rows = merged[merged['_merge'] == 'left_only']
#     #
#     # # 去除_merge列
#     # different_rows = different_rows.drop('_merge', axis=1)
#
#     # 将不同的行合并到df2中
#     # today_trade_df = pd.concat([today_trade_df, different_rows], ignore_index=True)
#
#     else:
#         # 使用 merge 和 indicator 来标记新增数据，找出不同行
#         merged_df = today_trade_df.merge(
#             existing_df,
#             # today_trade_df,
#             on=['组合名称', '代码', '名称', '操作', '最新价', '当前比例%', '新比例%', '时间'],
#             how='left',
#             indicator=True,
#             suffixes=('_new', '_old') # 明确指定后缀
#         )
#         # 仅保留在 today_trade_df 中存在但不在 existing_df 中的数据
#         new_data = merged_df[merged_df['_merge'] == 'left_only'].drop(columns=['_merge'])
#
#     if not new_data.empty:
#         logger.info(f'已存在的数据：\n {existing_df}')
#         logger.info(f'新增调仓：\n {new_data}')
#         # 合并新旧数据并去重
#         combined_df = pd.concat([existing_df, new_data], ignore_index=True)
#         combined_df.drop_duplicates(
#             # subset=['组合名称', '代码', '名称', '操作', '最新价', '当前比例%', '新比例%', '时间'],
#             subset=['组合名称', '代码', '名称', '操作', '最新价', '当前比例%', '新比例%', '时间'],
#             keep='first',
#             inplace=True
#         )
#         combined_df['时间'] = pd.to_datetime(combined_df['时间'])
#         combined_df.sort_values(by='时间', ascending=True, inplace=True)
#         logger.info(f'合并新旧数据：\n {combined_df}')
#
#         # 清空昨天的数据
#         clear_sheet(ETF_Combination_TODAY_ADJUSTMENT_FILE, sheet_name=sheet_name)  # 如果不需要清空，可以注释掉
#         save_to_excel(today_trades_without_sc_df, ETF_Combination_TODAY_ADJUSTMENT_FILE, sheet_name=sheet_name, index=False)
#
#         send_notification(f"今天有新调仓，{sheet_name}")
#         send_email(f'{sheet_name}策略调仓', combined_df.to_string(), '2695418206@qq.com')
#
#         create_flag_file(OPRATION_RECORD_DONE_FILE)
#     else:
#         logger.info("没有新增调仓数据")
async def strategy_main():
    strategy_id_to_name = Strategy_id_to_name
    strategy_ids = Strategy_ids
    all_today_trades_info = []
    all_latest_trade_info = []

    logger.info("开始处理策略调仓信息")
    for strategy_id in strategy_ids:
        combination_name = Strategy_id_to_name.get(strategy_id, '未知策略')
        latest_trade_info, trade_date = get_latest_position_and_trade(strategy_id)
        if latest_trade_info:
            all_latest_trade_info.extend(latest_trade_info)
            today_trades = [trade for trade in latest_trade_info if trade['时间'] == datetime.datetime.now().date().strftime('%Y%m%d')]
            all_today_trades_info.extend(today_trades)

    # 过滤掉创业板股票的交易信息
    all_latest_trade_info_without_sc = [trade for trade in all_latest_trade_info if trade['市场'] not in ['创业板', '科创板']]
    all_today_trades_info_without_sc = [trade for trade in all_today_trades_info if trade['市场'] not in ['创业板', '科创板']]
    # logger.info("去掉参考价大于30的")
    # all_today_trades_info_without_sc = [trade for trade in all_today_trades_info_without_sc if trade['参考价'] < 30]

    today_trades_without_sc_df = pd.DataFrame(all_today_trades_info_without_sc)

    file_path = STRATEGY_TODAY_ADJUSTMENT_FILE

    # 清空昨天的数据
    clear_sheet(file_path, '策略今天调仓')  # 如果不需要清空，可以注释掉

    # 保存今天的数据，即使为空也会覆盖昨天的数据
    save_to_excel(today_trades_without_sc_df, file_path, sheet_name='策略今天调仓', index=False)
    logger.info(f'今日调仓：\n {today_trades_without_sc_df}')

    if not today_trades_without_sc_df.empty:
        # 发送通知
        send_notification("今日有新的调仓操作！策略")
        send_email('股票策略调仓', today_trades_without_sc_df.to_string(), '2695418206@qq.com')
        # 创建标志文件
        with open(f"{OPRATION_RECORD_DONE_FILE}", "w") as f:
            f.write("策略调仓已完成")
            logger.info("创建标志文件成功")
    else:
        logger.info("未发送通知: 策略今天有调仓，但是是非沪深股票或无新增调仓")

    logger.info("策略调仓信息处理完成")

if __name__ == '__main__':
    strategy_id_to_name = Strategy_id_to_name
    import asyncio

    asyncio.run(strategy_main())

    # try:
    #     scheduler = Scheduler(interval=0.25,
    #                           callback=main,
    #                           start_time=dt_time(9, 29),
    #                           end_time=dt_time(10, 33))
    #     scheduler.start()
    # except Exception as e:
    #     logger.error(f"调度器启动失败: {e}", exc_info=True)

