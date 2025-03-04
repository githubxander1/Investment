import asyncio
import datetime
import os

import openpyxl
import pandas as pd
import requests

from others.量化投资.THS.自动化交易_同花顺.config.settings import ETF_ids, ETF_ids_to_name, \
    ETF_ADJUSTMENT_LOG_FILE, Combination_ids, \
    Combination_ids_to_name, ETF_Combination_TODAY_ADJUSTMENT_FILE
from others.量化投资.THS.自动化交易_同花顺.utils.determine_market import determine_market
from others.量化投资.THS.自动化交易_同花顺.utils.logger import setup_logger
from others.量化投资.THS.自动化交易_同花顺.utils.notification import send_notification

logger = setup_logger(ETF_ADJUSTMENT_LOG_FILE)

def fetch_and_extract_data(portfolio_id, is_etf=True):
    url = "https://t.10jqka.com.cn/portfolio/post/v2/get_relocate_post_list"
    headers = {
        "Host": "t.10jqka.com.cn",
        "Connection": "keep-alive",
        "Accept": "application/json, text/plain, */*",
        "User-Agent": "Mozilla/5.0 (Linux; Android 10; Redmi Note 7 Pro Build/QKQ1.190915.002; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/87.0.4280.101 Mobile Safari/537.36 Hexin_Gphone/11.16.10 (Royal Flush) hxtheme/1 innerversion/G037.08.980.1.32 followPhoneSystemTheme/1 userid/641926488 getHXAPPAccessibilityMode/0 hxNewFont/1 isVip/0 getHXAPPFontSetting/normal getHXAPPAdaptOldSetting/0",
        "Content-Type": "application/x-www-form-urlencoded",
        "X-Requested-With": "com.hexin.plat.android",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7",
        # "Cookie": 'user=MDptb182NDE5MjY0ODg6Ok5vbmU6NTAwOjY1MTkyNjQ4ODo3LDExMTExMTExMTExLDQwOzQ0LDExLDQwOzYsMSw0MDs1LDEsNDA7MSwxMDEsNDA7MiwxLDQwOzMsMSw0MDs1LDEsNDA7OCwwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMSw0MDsxMDIsMSw0MDoyNzo6OjY0MTkyNjQ4ODoxNzM3MzM4ODA5Ojo6MTY1ODE0Mjc4MDoyNjc4NDAwOjA6MTJiMmY0NGE2ODgxYjg0Nzc1YzY2MzM2MGM2NGUxZjMwOjox; userid=641926488; u_name=mo_641926488; escapename=mo_641926488; ticket=ee119caec220dd3e984ad47c01216b5f; user_status=0; IFUserCookieKey={"escapename":"mo_641926488","userid":"641926488"}; hxmPid=free_stock_159866.dstx; v=A2CljHMxEwZDm68CVQqtG88ZM2UyaUQz5k2YN9pxLHsO1Q9fgnkUwzZdaMQp'
        "Cookie": "IFUserCookieKey={\"escapename\":\"mo_641926488\",\"userid\":\"641926488\"}; user=MDptb182NDE5MjY0ODg6Ok5vbmU6NTAwOjY1MTkyNjQ4ODo3LDExMTExMTExMTExLDQwOzQ0LDExLDQwOzYsMSw0MDs1LDEsNDA7MSwxMDEsNDA7MiwxLDQwOzMsMSw0MDs1LDEsNDA7OCwwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMSw0MDsxMDIsMSw0MDoyNzo6OjY0MTkyNjQ4ODoxNzM5NzU2NTIyOjo6MTY1ODE0Mjc4MDoyNjc4NDAwOjA6MTkzMGZkYjc2OWQ2ZTE5OTI0MWNkYWVlZGE4YmFiYzA5Ojox; userid=641926488; u_name=mo_641926488; escapename=mo_641926488; ticket=3c97544e804cd21dbbfc935c45b489fa; user_status=0; hxmPid=ann_50844723; v=AyHkTzoW4ka_Bk6yYR38qIZGMuY7zpXAv0I51IP2HSiH6k4cyx6lkE-SSbQQ"

    }
    params = {"id": portfolio_id, "dynamic_id": 0}
    try:
        response = requests.get(url, params=params, headers=headers)
        response.raise_for_status()
        data = response.json()
    except requests.RequestException as e:
        logger.error(f"请求出错 (ID: {portfolio_id}): {e}")
        return []

    newest_post_infos = []

    if data is None:
        logger.info(f"ID: {portfolio_id} 无数据")
        return newest_post_infos

    data = data.get('data', None)
    if data is None:
        logger.info(f"ID: {portfolio_id} 无数据")
        return newest_post_infos
        # 如果 data 是列表，遍历每个元素
    if isinstance(data, list):
        for item in data:
            createAt = item.get('createAt', None)
            relocateList = item.get('relocateList', None)

            if relocateList:
                for infos in relocateList:
                    name = infos.get('name', None)
                    code = infos.get('code', None)
                    if '***' in name:
                        logger.warning(f"未订阅或股票名称显示异常 -组合id:{portfolio_id} 股票代码: {code}, 时间: {createAt}")
                        continue

                    # 计算操作类型
                    current_ratio = infos.get('currentRatio', 0)
                    new_ratio = infos.get('newRatio', 0)
                    operation = '买入' if new_ratio > current_ratio else '卖出'

                    market = determine_market(code)
                    # 获取组合名称
                    if is_etf:
                        combination_name = ETF_ids_to_name.get(portfolio_id, '未知ETF组合')
                    else:
                        combination_name = Combination_ids_to_name.get(portfolio_id, '未知股票组合')

                    newest_post = {
                        '组合名称': combination_name,
                        '代码':  str(code).zfill(6),  # 提前统一格式
                        '股票名称': name,
                        '市场': market,
                        '操作': operation,
                        '最新价': infos.get('finalPrice'),
                        '当前比例%': round(current_ratio * 100, 2),
                        '新比例%': round(new_ratio * 100, 2),
                        '时间': createAt
                    }
                    newest_post_infos.append(newest_post)
    else:
        logger.warning(f"ID: {portfolio_id} 数据格式异常，期望列表，实际为 {type(data)}")

    return newest_post_infos

def save_to_excel(df, filepath, sheet_name, index=False):
    """将DataFrame保存到Excel文件中"""
    try:
        # 检查文件是否存在
        try:
            with pd.ExcelFile(filepath) as _:
                # 文件存在，追加模式
                with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=index)
        except FileNotFoundError:
            # 文件不存在，创建新文件
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
        logger.info(f"成功保存数据到文件: {filepath}, 表名称: {sheet_name}")
    except Exception as e:
        logger.error(f"保存数据到文件失败: {e}")

def clear_sheet(filepath, sheet_name):
    """清空指定Excel文件中的指定表格"""
    try:
        wb = openpyxl.load_workbook(filepath)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
            wb.save(filepath)
            logger.info(f"成功清空表格: {sheet_name}")
        else:
            logger.debug(f"表格 {sheet_name} 不存在，无需清空")  # 降低日志级别 ✅
    except FileNotFoundError:
        logger.debug(f"文件 {filepath} 不存在，无需清空")  # 降低日志级别 ✅
    except Exception as e:
        logger.error(f"清空表格失败: {e}")

def create_flag_file(flag_file):
    """创建标志文件"""
    try:
        with open(flag_file, "w") as f:
            f.write("已创建标志文件")
        logger.info(f"创建标志文件: {flag_file}")
    except Exception as e:
        logger.error(f"创建标志文件失败: {e}")

def process_today_trades(ids, is_etf=True):
    """处理今天调仓数据"""
    today_records = []
    today = datetime.date.today().strftime('%Y-%m-%d')

    for portfolio_id in ids:
        extracted_result = fetch_and_extract_data(portfolio_id, is_etf)
        for item in extracted_result:
            create_at = item['时间']
            date_part = create_at.split()[0]
            if date_part == today:
                today_records.append(item)

    if not today_records:
        logger.info(f"所选{'ETF' if is_etf else '股票组合'}今天无调仓")
        return None

    today_trade_df = pd.DataFrame(today_records)
    # 将 '时间' 列转换为 datetime 类型
    today_trade_df['时间'] = pd.to_datetime(today_trade_df['时间'])

       # 去掉创业板和科创板
    # today_trade_df = today_trade_df[~today_trade_df['代码'].str.startswith('60') | today_trade_df['代码'].str.startswith('30')]
    today_trade_df = today_trade_df[~today_trade_df['市场'].isin(['创业板', '科创板'])]


    return today_trade_df
async def check_new_data(existing_df, today_trade_df, sheet_name):
    """优化后的数据合并函数

    对比新旧数据
    找出新增
    附加补充新增数据到旧表

    Args:
        existing_df: 已存在的历史数据DataFrame
        today_trade_df: 当天获取的新数据DataFrame
        sheet_name: 要操作的Excel表名称

    Returns:
        pd.DataFrame: 合并后的完整数据集
    """
    try:
        # 预处理时间字段（精确到分钟）
        today_trade_df = today_trade_df.copy()
        today_trade_df['时间'] = pd.to_datetime(today_trade_df['时间']).dt.floor('Min')

        # 生成唯一标识（组合名称+代码+操作+新比例+时分）
        def create_unique_id(row):
            try:
                # 统一股票代码为6位格式（自动补零）
                formatted_code = str(row['代码']).zfill(6)  # 关键修复点

                ratio = "{:.2f}".format(float(row['新比例%']))
                time_str = row['时间'].strftime('%H%M%S')
                return f"{row['组合名称']}_{formatted_code}_{row['操作']}_{ratio}_{time_str}"
            except Exception as e:
                logger.error(f"生成唯一标识失败: {str(e)}")
                return "error_id"


        today_trade_df['strict_id'] = today_trade_df.apply(create_unique_id, axis=1)

        # 处理空历史数据情况
        if existing_df.empty:
            # 首次保存时添加时间排序
            sorted_df = today_trade_df.sort_values('时间', ascending=True)
            save_to_excel(sorted_df, ETF_Combination_TODAY_ADJUSTMENT_FILE, sheet_name)
            # # 原代码中的通知部分：
            # send_notification(f"首次发现调仓，{sheet_name}")
            #
            # # 修改为：
            # notification_msg = f"{sheet_name}调仓操作\n" + "\n".join(
            #     [f"{row['组合名称']} {row['操作']} {row['代码']}"
            #      for _, row in new_data.iterrows()])
            # send_notification(notification_msg)

            return sorted_df

        # 标准化历史数据格式
        existing_df = existing_df.copy()
        existing_df['strict_id'] = existing_df.apply(create_unique_id, axis=1)  # 统一列名

        # 找出新增数据（修复查询条件）
        mask = ~today_trade_df['strict_id'].isin(existing_df['strict_id'])
        new_data = today_trade_df[mask].copy()

        # 合并数据并去重
        if not new_data.empty:
            # 合并前再次去重（防止多次运行产生重复）
            updated_df = pd.concat([existing_df, new_data], ignore_index=True)
            updated_df = updated_df.drop_duplicates(subset=['strict_id'], keep='last')

            # 按精确时间排序（秒级精度）
            sorted_df = updated_df.sort_values('时间', ascending=True).reset_index(drop=True)

            # 原子写入操作
            temp_file = ETF_Combination_TODAY_ADJUSTMENT_FILE.replace('.xlsx', '_temp.xlsx')
            save_to_excel(sorted_df, temp_file, sheet_name)
            # send_notification(f"新调仓，{sheet_name}")
            # # 原代码中的通知部分：
            # send_notification(f"首次发现调仓，{sheet_name}")

            # # 修改为：{sheet_name}操作
            notification_msg = f"\n> " + "\n".join(
                [f"{row['组合名称']} {row['操作']} {row['股票名称']} {row['新比例%']}% {row['最新价']} \n{row['时间']}"
                 for _, row in new_data.iterrows()])
            send_notification(notification_msg)

            if os.path.exists(temp_file):
                os.replace(temp_file, ETF_Combination_TODAY_ADJUSTMENT_FILE)

            logger.info(f"新增{len(new_data)}条唯一调仓记录")
        else:
            logger.info("没有新增调仓数据")
            sorted_df = existing_df

        return sorted_df.drop(columns=['strict_id'], errors='ignore')

    except Exception as e:
        logger.error(f"数据处理失败: {str(e)}")
        # 保留临时文件供调试
        if 'temp_file' in locals() and os.path.exists(temp_file):
            logger.error(f"临时文件保留在: {temp_file}")
        return existing_df


async def ETF_Combination_main():
    logger.info("开始处理ETF和股票组合调仓信息")

    # 处理 ETF 组合
    etf_today_trade_df = process_today_trades(ETF_ids, is_etf=True)
    # print('ETF今日调仓：\n')
    # logger.info(etf_today_trade_df)
    if etf_today_trade_df is not None:
        logger.info(f'ETF今日调仓：\n {etf_today_trade_df}')
        try:
            existing_etf_df = pd.read_excel(ETF_Combination_TODAY_ADJUSTMENT_FILE, sheet_name='ETF最新调仓')
            # print('已存在数据ETF：')
            # print(existing_etf_df)
        except FileNotFoundError:
            existing_etf_df = pd.DataFrame()
            logger.info("ETF Excel文件不存在，创建新文件")
        except ValueError:
            existing_etf_df = pd.DataFrame()
            logger.info("ETF工作表不存在，创建新工作表")

        await check_new_data(existing_etf_df, etf_today_trade_df, sheet_name='ETF最新调仓')
    else:
        logger.info("今天没有新的ETF调仓操作")

    # 处理股票组合
    stock_today_trade_df = process_today_trades(Combination_ids, is_etf=False)
    # print('股票组合今日调仓：\n')
    # print(stock_today_trade_df)
    if stock_today_trade_df is not None:
        logger.info(f'股票组合今日调仓：\n {stock_today_trade_df}')
        try:
            existing_stock_df = pd.read_excel(ETF_Combination_TODAY_ADJUSTMENT_FILE, sheet_name='股票组合最新调仓')
            # print('已存在数据股票：')
            # print(existing_stock_df)
        except FileNotFoundError:
            existing_stock_df = pd.DataFrame(columns=stock_today_trade_df.columns)  # ✅ 带列名
            logger.info("股票组合 Excel文件不存在，创建新文件")
        except ValueError:
            existing_stock_df = pd.DataFrame(columns=stock_today_trade_df.columns)  # ✅ 带列名
            logger.info("股票组合工作表不存在，创建新工作表")

        await check_new_data(existing_stock_df, stock_today_trade_df, sheet_name='股票组合最新调仓')
    else:
        logger.info("今天没有新的股票组合调仓操作")



if __name__ == "__main__":
    asyncio.run(ETF_Combination_main())
