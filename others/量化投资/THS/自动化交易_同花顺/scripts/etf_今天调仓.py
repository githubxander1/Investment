import asyncio
import datetime

import openpyxl
import pandas as pd
import requests

from others.量化投资.THS.自动化交易_同花顺.config.settings import ETF_ids, ETF_ids_to_name, \
    ETF_TODAY_ADJUSTMENT_FILE, ETF_ADJUSTMENT_LOG_FILE, OPRATION_RECORD_DONE_FILE
from others.量化投资.THS.自动化交易_同花顺.utils.logger import setup_logger
from others.量化投资.THS.自动化交易_同花顺.utils.notification import send_notification, send_email

logger = setup_logger(ETF_ADJUSTMENT_LOG_FILE)

def send_request(id):
    url = 'https://t.10jqka.com.cn/portfolio/post/v2/get_newest_relocate_post'
    params = {
        'id': id
    }
    headers = {
        'Host': 't.10jqka.com.cn',
        'Connection': 'keep-alive',
        'Accept': 'application/json, text/plain, */*',
        'User-Agent': 'Mozilla/5.0 (Linux; Android 10; Redmi Note 7 Pro Build/QKQ1.190915.002; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/87.0.4280.101 Mobile Safari/537.36 Hexin_Gphone/11.19.03 (Royal Flush) hxtheme/1 innerversion/G037.08.990.1.32 followPhoneSystemTheme/1 userid/641926488 getHXAPPAccessibilityMode/0 hxNewFont/1 isVip/0 getHXAPPFontSetting/normal getHXAPPAdaptOldSetting/0',
        'Content-Type': 'application/x-www-form-urlencoded',
        'X-Requested-With': 'com.hexin.plat.android',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Dest': 'empty',
        'Referer': 'https://t.10jqka.com.cn/pkgfront/tgService.html?type=portfolio&id=29617',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
        'Cookie': 'user=MDptb182NDE5MjY0ODg6Ok5vbmU6NTAwOjY1MTkyNjQ4ODo3LDExMTExMTExMTExLDQwOzQ4ODA5Ojo6MTY1ODE0Mjc4MDoyNjc4NDAwOjA6MTJiMmY0NGE2ODgxYjg0Nzc1YzY2MzM2MGM2NGUxZjMwOjox; userid=641926488; u_name=mo_641926488; escapename=mo_641926488; ticket=ee119caec220dd3e984ad47c01216b5f; user_status=0; IFUserCookieKey={"escapename":"mo_641926488","userid":"641926488"}; hxmPid=hqMarketPkgVersionControl; v=AxLXmrX7oTmTPd1F4fgfxZGDYdP0Ixa0SCcK4dxrPkWw771JxLNmzRi3WuOv'
    }
    try:
        response = requests.get(url, params=params, headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        logger.error(f"请求出错 (ID: {id}): {e}")
        return None

def extract_result(data, id):
    newest_post_infos = []

    if data is None:
        logger.info(f"ID: {id} 无数据")
        return newest_post_infos

    data = data.get('data', None)
    if data is None:
        logger.info(f"ID: {id} 无数据")
        return newest_post_infos

    content = data.get('content', None)
    createAt = data.get('createAt', None)
    dynamicId = data.get('dynamicId', None)

    relocateList = data.get('relocateList', None)

    for infos in relocateList:
        name = infos.get('name', None)
        code = infos.get('code', None)
        # 检查股票名称是否为“匿名”
        if '***' in name:
            logger.warning(f"未订阅或股票名称显示异常 -组合id:{id} 股票代码: {code}, 时间: {createAt}")
            continue

        newest_post = {
            'ETF组合': ETF_ids_to_name.get(id, '未知'),
            '代码': infos.get('code'),
            '名称': infos.get('name'),
            '最新价': infos.get('finalPrice'),
            '当前比例%': round(infos.get('currentRatio') * 100,2),
            '新比例%': round(infos.get('newRatio') * 100,2),
            # '说明': content,
            '时间': createAt
        }
        newest_post_infos.append(newest_post)

    return newest_post_infos

def save_to_excel(df, filename, sheet_name, index=False):
    """将DataFrame保存到Excel文件中"""
    try:
        # 检查文件是否存在
        try:
            with pd.ExcelFile(filename) as _:
                # 文件存在，追加模式
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=index)
        except FileNotFoundError:
            # 文件不存在，创建新文件
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
        logger.info(f"成功保存数据到文件: {filename}, 表名称: {sheet_name}")
    except Exception as e:
        logger.error(f"保存数据到文件失败: {e}")

def clear_sheet(filename, sheet_name):
    """清空指定Excel文件中的指定表格"""
    try:
        # 检查文件是否存在
        try:
            wb = openpyxl.load_workbook(filename)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 删除所有行
                ws.delete_rows(1, ws.max_row)
                wb.save(filename)
                logger.info(f"成功清空表格: {sheet_name} 文件: {filename}")
            else:
                logger.warning(f"表格 {sheet_name} 不存在于文件: {filename}")
        except FileNotFoundError:
            logger.warning(f"文件 {filename} 不存在，无需清空")
    except Exception as e:
        logger.error(f"清空表格失败: {e}")

def create_flag_file(flag_file):
    """创建标志文件"""
    try:
        with open(flag_file, "w") as f:
            f.write("已创建标志文件")
        logger.info(f"创建标志文件: {flag_file}")
    except Exception as e:
        logger.error(f"创建标志文件失败: {e}")

def process_today_trades(ids):
    """处理今天调仓数据"""
    all_records = []
    today = datetime.date.today().strftime('%Y-%m-%d')

    for portfolio_id in ids:
        data = send_request(portfolio_id)
        if data:
            extracted_result = extract_result(data, portfolio_id)
            for item in extracted_result:
                create_at = item['时间']
                date_part = create_at.split()[0]
                if date_part == today:
                    all_records.append(item)

    if not all_records:
        logger.info("所选ETF今天无调仓")
        return None

    today_trade_df = pd.DataFrame(all_records)
    # 将 '时间' 列转换为 datetime 类型
    today_trade_df['时间'] = pd.to_datetime(today_trade_df['时间'])
    return today_trade_df

def check_new_data(existing_df, today_trade_df):
    # 确保 '代码' 列的数据类型一致
    today_trade_df['代码'] = today_trade_df['代码'].astype(str)
    existing_df['代码'] = existing_df['代码'].astype(str)

    # 确保 '时间' 列的数据类型一致
    today_trade_df['时间'] = pd.to_datetime(today_trade_df['时间'])
    existing_df['时间'] = pd.to_datetime(existing_df['时间'])

    if not existing_df.empty:
        new_data = today_trade_df.merge(existing_df, how='outer', indicator=True).loc[
            lambda x: x['_merge'] == 'left_only'].drop(columns='_merge')
    else:
        new_data = today_trade_df.copy()

    if not new_data.empty:
        logger.info(f'新增调仓：\n {new_data}')
        combined_df = pd.concat([new_data, today_trade_df], ignore_index=True)
        combined_df.drop_duplicates(subset=None, keep='first', inplace=True)
        combined_df['时间'] = pd.to_datetime(combined_df['时间'])
        combined_df.sort_values(by='时间', ascending=True, inplace=True)
        logger.info(f'合并新旧数据：\n {combined_df}')
        save_to_excel(combined_df, ETF_TODAY_ADJUSTMENT_FILE, sheet_name='ETF最新调仓', index=False)

        send_notification("今天有新调仓，ETF")
        send_email('ETF策略调仓', combined_df.to_string(), '2695418206@qq.com')

        create_flag_file(OPRATION_RECORD_DONE_FILE)
    else:
        logger.info("没有新增调仓数据")

async def ETF_main():
    logger.info("开始处理ETF调仓信息")
    today_trade_df = process_today_trades(ETF_ids)
    # print(today_trade_df)

    if today_trade_df is not None:
        logger.info(f'今日调仓：\n {today_trade_df}')
        try:
            existing_df = pd.read_excel(ETF_TODAY_ADJUSTMENT_FILE, sheet_name='ETF最新调仓')
        except FileNotFoundError:
            existing_df = pd.DataFrame()
            # print(existing_df)
            existing_df.to_excel(ETF_TODAY_ADJUSTMENT_FILE, sheet_name='ETF最新调仓', index=False)
            logger.info("Excel文件不存在，创建新文件")

        check_new_data(existing_df, today_trade_df)
    else:
        logger.info("今天没有新的调仓操作")

if __name__ == "__main__":
    asyncio.run(ETF_main())
