# utils/excel_handler.py

import json

import openpyxl

from others.测试相关.api_mind.utils.logger import logger


def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    headers = [cell.value for cell in sheet[6]]  # 第六行是标题行
    for row in sheet.iter_rows(min_row=7, values_only=True):  # 从第七行开始是测试用例数据
        row_dict = dict(zip(headers, row))
        # 尝试将预期结果解析为字典
        try:
            # 将字符串中的单引号替换为双引号，因为 JSON 标准使用双引号
            row_dict["预期结果"] = row_dict["预期结果"].replace("'", '"')
            # 使用 json 模块的 loads 方法将字符串解析为字典
            row_dict["预期结果"]= json.loads(row_dict["预期结果"])
            # row_dict["预期结果"] = json.loads(row_dict["预期结果"])
            # print(type(row_dict["预期结果"]))
        except (json.JSONDecodeError, TypeError):
            logger.error(f"Failed to parse expected result: {row_dict['预期结果']}")
            row_dict["预期结果"] = {}
        data.append(row_dict)
    return data

def write_excel(file_path, row_index, column_index, value):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    sheet.cell(row=row_index, column=column_index, value=value)
    wb.save(file_path)

# if __name__ == '__main__':
#     pprint(read_excel('../testdata/rename.xlsx'))
