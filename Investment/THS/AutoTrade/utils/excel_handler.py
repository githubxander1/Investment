from datetime import datetime
import os
from pprint import pprint

import openpyxl
import pandas as pd

from Investment.THS.AutoTrade.config.settings import Strategy_portfolio_today, Account_holding_stockes_info_file
# from Investment.THS.AutoTrade.scripts.process_stocks_to_operate_data import read_operation_history, \
#     write_operation_history
from Investment.THS.AutoTrade.utils.notification import send_notification
from Investment.THS.AutoTrade.utils.scheduler import logger
from Investment.THS.AutoTrade.utils.time_utils import normalize_time


def create_empty_excel(file_path, sheet_name):
    if not os.path.exists(file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            pd.DataFrame(columns=['组合名称', '代码', '操作', '新比例%', '时间']).to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info(f"创建空Excel文件: {file_path}, 表名称: {sheet_name}")

def read_excel(file_path, sheet_name):
    try:
        # 读取Excel文件
        if os.path.exists(file_path):
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            # pprint(f"成功读取文件: {file_path}, 表名称: {sheet_name}")
            return df
        else:
            logger.warning(f"文件 {file_path} 不存在，返回空DataFrame")
            return pd.DataFrame()
    except Exception as e:
        logger.error(f"读取文件失败: {e}")
        return pd.DataFrame()

def read_portfolio_record_history(file_path):
    today = normalize_time(datetime.now().strftime('%Y-%m-%d'))
    if os.path.exists(file_path):
        try:
            with pd.ExcelFile(file_path, engine='openpyxl') as operation_history_xlsx:
                if today in operation_history_xlsx.sheet_names:
                    portfolio_record_history_df = pd.read_excel(operation_history_xlsx, sheet_name=today)
                    # 去重处理
                    portfolio_record_history_df.drop_duplicates(subset=['标的名称', '操作', '新比例%', '时间'], inplace=True)
                    logger.info(f"读取去重后的操作历史文件完成\n{portfolio_record_history_df}")
                else:
                    portfolio_record_history_df = pd.DataFrame(columns=["名称","操作","标的名称","代码","最新价","新比例%","市场","时间"])
                    logger.warning(f"历史文件{portfolio_record_history_df},表名称: {today}不存在")
        except Exception as e:
            logger.error(f"读取操作历史文件失败: {e}", exc_info=True)
            portfolio_record_history_df = pd.DataFrame(columns=['标的名称', '操作', '状态', '信息', '时间'])
    else:
        portfolio_record_history_df = pd.DataFrame(columns=["名称","操作","标的名称","代码","最新价","新比例%","市场","时间"])

    return portfolio_record_history_df

def save_to_excel(df, filename, sheet_name, index=False):
    """追加保存DataFrame到Excel文件"""
    try:
        # 调试：打印要保存的数据
        print(f"即将保存的数据:\n{df}")

        # 检查文件是否存在
        if os.path.exists(filename):
            # 文件存在，读取现有 sheet 数据并追加
            with pd.ExcelFile(filename, engine='openpyxl') as xls:
                if sheet_name in xls.sheet_names:
                    existing_df = pd.read_excel(xls, sheet_name=sheet_name)
                    combined_df = pd.concat([existing_df, df], ignore_index=True)
                else:
                    combined_df = df

            # 写回整个 DataFrame 到指定 sheet
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                combined_df.to_excel(writer, sheet_name=sheet_name, index=index)
            logger.info(f"✅ 成功追加数据到Excel文件: {filename}, 表名称: {sheet_name}")
        else:
            # 文件不存在，创建新文件
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
            logger.info(f"✅ 创建并保存数据到Excel文件: {filename}, 表名称: {sheet_name}")

    except Exception as e:
        logger.error(f"❌ 保存数据到Excel文件失败: {e}", exc_info=True)


def verify_excel_file_update(filename, sheet_name, expected_rows):
    if not os.path.exists(filename):
        return False
    try:
        df = pd.read_excel(filename, sheet_name=sheet_name)
        if len(df) > expected_rows:
            logger.info("✅ 文件内容已更新")
            return True
        else:
            logger.warning("⚠️ 文件未新增数据，请检查写入逻辑")
            return False
    except Exception as e:
        logger.error(f"验证文件更新失败: {e}")
        return False

def process_excel_files(ths_page, file_paths, operation_history_file, holding_stock_file):
    for file_path in file_paths:
        logger.info(f"检测到文件更新，即将进行操作的文件路径: {file_path}")
        if not os.path.exists(file_path):
            logger.warning(f"文件不存在: {file_path}")
            continue
        try:
            # 读取要处理的文件
            df = pd.read_csv(file_path)
            for index, row in df.iterrows():
                stock_name = row['标的名称']
                operation = row['操作']
                time = row['时间']
                price = int(row['最新价'])
                new_ratio = float(row['新比例%'])
                # logger.info(f"要处理的信息:  {operation} {stock_name} {new_ratio}")
                logger.info(f"要处理的信息:  {operation} {stock_name} {price}")

                # 读取最新的操作历史记录
                operation_history_df = read_operation_history(operation_history_file)

                # 检查标的名称和操作,新比例是否已经存在于操作历史中
                if not operation_history_df.empty:
                    existing_operations = operation_history_df[(operation_history_df['标的名称'] == stock_name) & (operation_history_df['新比例%'] == new_ratio)]
                    if not existing_operations.empty:
                        logger.info(f"{stock_name} 和操作 {operation} {new_ratio}已经操作过，跳过")
                        continue

                #定义操作股数：如果operation=“买入”，则买入数量为4000除以价格，结果取整；如果为卖出，如果new_ratio不为0，则卖出半仓，如果为0，则卖出全部
                holding_stock_df = pd.read_excel(Account_holding_stockes_info_file, sheet_name="持仓数据")
                print(holding_stock_df)
                holding_stock = holding_stock_df[holding_stock_df['标的名称'] == stock_name]
                if not holding_stock.empty:
                    available = int(holding_stock['持仓/可用'].str.split('/').str[1].iloc[0])
                    print(f"持仓数量: {available}")
                real_price = ths_page.get_real_price()
                volume = int(4000 / real_price) if operation == "买入" else int(4000 / real_price / 2) if new_ratio != '0' else available

                #开始操作
                status, info = ths_page.sell_stock(stock_name, volume) if operation == '卖出' else ths_page.buy_stock(stock_name, volume)
                # logger.info(f"{operation} {stock_name} {'成功' if status else '失败'} {info}")
                logger.info(f"{operation} {stock_name} {info}")
                send_notification(f"{operation} {stock_name} {info}")
                logger.info(f"发送通知成功")

                # 写入操作历史
                operate_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                new_operation_history = pd.DataFrame([{
                    '标的名称': stock_name,
                    '操作': operation,
                    '新比例%': new_ratio,
                    '状态': status,
                    '信息': info,
                    '时间': operate_time
                }])
                write_operation_history(new_operation_history)
        except Exception as e:
            logger.error(f"处理文件 {file_path} 失败: {e}", exc_info=True)

# def clear_sheet(filename, sheet_name):
#     """清空指定Excel文件中的指定表格"""
#     try:
#         # 检查文件是否存在
#         if os.path.exists(filename):
#             wb = openpyxl.load_workbook(filename)
#             if sheet_name in wb.sheetnames:
#                 ws = wb[sheet_name]
#                 # 删除所有行
#                 ws.delete_rows(1, ws.max_row)
#                 wb.save(filename)
#                 pprint(f"成功清空表格: {sheet_name} 文件: {filename}")
#             else:
#                 logger.warning(f"表格 {sheet_name} 不存在于文件: {filename}")
#         else:
#             logger.warning(f"文件 {filename} 不存在，无需清空")
#     except Exception as e:
#         logger.error(f"清空表格失败: {e}")
def clear_csv(filename):
    try:
        if os.path.exists(filename):
            with open(filename, 'w', newline='') as f:
                f.truncate()
                logger.info(f"成功清空文件: {filename}")
        else:
            logger.warning(f"文件 {filename} 不存在，无需清空")
    except  Exception as e:
        logger.error(f"清空文件失败: {e}")


if __name__ == '__main__':
    # read_portfolio_record_history(Strategy_portfolio_today)
    df = pd.DataFrame(columns=['标的名称', '操作', '新比例%'])
    save_to_excel(df,Strategy_portfolio_today,'今天',index=True)